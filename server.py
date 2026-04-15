import os
import secrets
from datetime import datetime, date, timedelta
from calendar import monthrange
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO
from flask import Flask, render_template, request, redirect, jsonify, session, url_for, send_file
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import text

app = Flask(__name__)
app.config['SECRET_KEY'] = 'dev-key-123'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///company.db'
db = SQLAlchemy(app)

class Organization(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    invite_code = db.Column(db.String(10), unique=True, nullable=True)
    invite_code_expires = db.Column(db.DateTime, nullable=True)
    test_data_loaded = db.Column(db.Boolean, default=False)
    users = db.relationship('User', backref='organization', lazy=True)
    tasks = db.relationship('Task', backref='organization', lazy=True)

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    first_name = db.Column(db.String(50))
    last_name = db.Column(db.String(50))
    login = db.Column(db.String(50))
    password = db.Column(db.String(50))
    hourly_rate = db.Column(db.Float, default=0)
    role = db.Column(db.String(10), default='user')
    permissions = db.Column(db.String(200), nullable=True)
    organization_id = db.Column(db.Integer, db.ForeignKey('organization.id'))
    __table_args__ = (db.UniqueConstraint('login', 'organization_id', name='_user_login_org_uc'),)

class Task(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(100))
    deadline = db.Column(db.Date)
    is_done = db.Column(db.Boolean, default=False)
    organization_id = db.Column(db.Integer, db.ForeignKey('organization.id'))
    users = db.relationship('User', secondary='task_assignees', backref='tasks')
    completions = db.relationship('TaskCompletion', backref='task', lazy=True)

class TaskCompletion(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    task_id = db.Column(db.Integer, db.ForeignKey('task.id'))
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    completed = db.Column(db.Boolean, default=False)
    __table_args__ = (db.UniqueConstraint('task_id', 'user_id'),)

task_assignees = db.Table('task_assignees',
    db.Column('user_id', db.Integer, db.ForeignKey('user.id')),
    db.Column('task_id', db.Integer, db.ForeignKey('task.id'))
)

class WorkSession(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    date = db.Column(db.Date, default=date.today)
    duration_minutes = db.Column(db.Integer, default=0)
    manual_adjustment = db.Column(db.Boolean, default=False)
    adjusted_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    adjust_reason = db.Column(db.String(200), nullable=True)

# ---------- Helpers ----------
def generate_invite_code():
    return secrets.token_hex(4).upper()

def get_current_organization():
    if 'org_id' in session:
        return Organization.query.get(session['org_id'])
    return None

def is_superadmin(user):
    return user is not None and user.role == 'admin' and user.permissions == '*'

def has_perm(user, perm):
    if user is None or user.role != 'admin':
        return False
    if user.permissions == '*':
        return True
    if user.permissions:
        return perm in user.permissions.split(',')
    return False

def get_perms_list(user):
    if user is None or user.role != 'admin':
        return []
    if user.permissions == '*':
        return ['tasks', 'profiles', 'org', 'superadmin']
    if user.permissions:
        return user.permissions.split(',')
    return []

def get_period_bounds(period, custom_start=None, custom_end=None):
    today = date.today()
    if period == 'week':
        start = today - timedelta(days=today.weekday())
        end = today
    elif period == 'month':
        start = today.replace(day=1)
        end = today
    elif period == 'quarter':
        q_start_month = ((today.month - 1) // 3) * 3 + 1
        start = today.replace(month=q_start_month, day=1)
        end = today
    elif period == 'year':
        start = today.replace(month=1, day=1)
        end = today
    elif period == 'alltime':
        start = date(2020, 1, 1)
        end = today
    elif period == 'custom' and custom_start and custom_end:
        try:
            start = datetime.strptime(custom_start, '%Y-%m-%d').date()
            end = datetime.strptime(custom_end, '%Y-%m-%d').date()
        except:
            start = today.replace(day=1)
            end = today
    else:
        start = today - timedelta(days=30)
        end = today
    return start, end

def run_migrations():
    with db.engine.connect() as conn:
        try:
            conn.execute(text('ALTER TABLE user ADD COLUMN permissions VARCHAR(200)'))
            conn.commit()
        except Exception:
            pass
        try:
            conn.execute(text('ALTER TABLE organization ADD COLUMN test_data_loaded BOOLEAN DEFAULT 0'))
            conn.commit()
        except Exception:
            pass

def _seed_employees_and_data(org):
    import random
    employees_data = [
        ('Александр', 'Петров',   'petrov',    'pass1', 350.0),
        ('Мария',     'Сидорова',  'sidorova',  'pass2', 420.0),
        ('Дмитрий',  'Козлов',    'kozlov',    'pass3', 280.0),
        ('Елена',    'Новикова',   'novikova',  'pass4', 510.0),
        ('Иван',     'Морозов',    'morozov',   'pass5', 390.0),
        ('Анна',     'Волкова',    'volkova',   'pass6', 460.0),
        ('Сергей',   'Зайцев',     'zaitsev',   'pass7', 320.0),
        ('Ольга',    'Лебедева',   'lebedeva',  'pass8', 575.0),
        ('Алексей',  'Семёнов',    'semenov',   'pass9', 245.0),
        ('Наталья',  'Орлова',     'orlova',    'pass10', 620.0),
    ]
    employees = []
    for fn, ln, login, pw, rate in employees_data:
        if User.query.filter_by(login=login, organization_id=org.id).first():
            continue
        u = User(first_name=fn, last_name=ln, login=login,
                 password=pw, hourly_rate=rate, role='user',
                 organization_id=org.id)
        db.session.add(u)
        employees.append(u)
    db.session.flush()

    today = date.today()
    start_date = today - timedelta(days=270)
    for emp in employees:
        current = start_date
        while current <= today:
            if current.weekday() < 5:
                variance = random.randint(-30, 30)
                minutes = max(300, 480 + variance)
                if random.random() < 0.1:
                    minutes = random.randint(60, 240)
                db.session.add(WorkSession(user_id=emp.id, date=current, duration_minutes=minutes))
            current += timedelta(days=1)

    tasks_done = [
        ('Разработка модуля авторизации',    date(2025, 11, 15)),
        ('Вёрстка главной страницы',         date(2025, 12, 1)),
        ('Тестирование API',                 date(2026, 1, 10)),
        ('Написание документации',           date(2026, 2, 5)),
        ('Оптимизация базы данных',          date(2026, 3, 20)),
    ]
    tasks_active = [
        ('Разработка мобильного приложения', today + timedelta(days=30)),
        ('Миграция на новый сервер',         today + timedelta(days=14)),
        ('Рефакторинг кода',                 today + timedelta(days=7)),
        ('Интеграция с CRM',                 today + timedelta(days=45)),
        ('Обучение новых сотрудников',       today + timedelta(days=60)),
    ]
    for title, deadline in tasks_done:
        t = Task(title=title, deadline=deadline, is_done=True, organization_id=org.id)
        db.session.add(t)
        db.session.flush()
        for emp in employees:
            t.users.append(emp)
            db.session.add(TaskCompletion(task_id=t.id, user_id=emp.id, completed=True))
    for title, deadline in tasks_active:
        t = Task(title=title, deadline=deadline, is_done=False, organization_id=org.id)
        db.session.add(t)
        db.session.flush()
        for emp in employees:
            t.users.append(emp)
            db.session.add(TaskCompletion(task_id=t.id, user_id=emp.id, completed=False))
    db.session.commit()
    return employees

# ---------- Auth ----------
@app.route('/register', methods=['GET', 'POST'])
def register_org():
    if request.method == 'POST':
        org_name = request.form['org_name']
        org = Organization(name=org_name)
        db.session.add(org)
        db.session.flush()
        admin = User(
            first_name=request.form['first_name'],
            last_name=request.form['last_name'],
            login=request.form['login'],
            password=request.form['password'],
            hourly_rate=0,
            role='admin',
            permissions='*',
            organization_id=org.id
        )
        db.session.add(admin)
        db.session.commit()
        session['user_id'] = admin.id
        session['org_id'] = org.id
        return render_template('register.html', created=True, org_id=org.id, org_name=org.name)
    return render_template('register.html', created=False)

@app.route('/')
def home():
    return redirect(url_for('login_page'))

@app.route('/login', methods=['GET', 'POST'])
def login_page():
    if request.method == 'POST':
        org_id = request.form.get('org_id')
        login = request.form['login']
        password = request.form['password']
        user = User.query.filter_by(login=login, password=password, organization_id=org_id).first()
        if user:
            session['user_id'] = user.id
            session['org_id'] = user.organization_id
            return redirect(url_for('dashboard'))
    return render_template('login.html')

@app.route('/dashboard')
def dashboard():
    if 'user_id' not in session:
        return redirect(url_for('login_page'))
    user = User.query.get(session['user_id'])
    org = Organization.query.get(session['org_id'])
    if user.role == 'admin':
        return redirect(url_for('admin_panel'))
    today = date.today()
    return render_template('user_dashboard.html', user=user, org=org, today=today)

@app.route('/admin')
def admin_panel():
    if 'user_id' not in session:
        return redirect(url_for('login_page'))
    admin = User.query.get(session['user_id'])
    if admin.role != 'admin':
        return redirect(url_for('dashboard'))
    org = Organization.query.get(session['org_id'])
    users = User.query.filter_by(organization_id=org.id, role='user').all()
    admins = User.query.filter_by(organization_id=org.id, role='admin').all()
    tasks = Task.query.filter_by(organization_id=org.id).all()
    now = datetime.utcnow()
    error = request.args.get('error')
    perms = get_perms_list(admin)
    return render_template('admin_dashboard.html', org=org, users=users, admins=admins,
                           tasks=tasks, now=now, error=error, current_admin=admin, perms=perms)

@app.route('/admin/update_org', methods=['POST'])
def update_org():
    org = get_current_organization()
    admin = User.query.get(session['user_id'])
    if not org or not has_perm(admin, 'org'):
        return redirect(url_for('admin_panel', error='Нет прав для изменения организации.'))
    org.name = request.form['name']
    admin.first_name = request.form.get('admin_first_name', admin.first_name)
    admin.last_name = request.form.get('admin_last_name', admin.last_name)
    if request.form.get('password'):
        admin.password = request.form['password']
    db.session.commit()
    return redirect(url_for('admin_panel'))

@app.route('/admin/seed_test_data', methods=['POST'])
def admin_seed_test_data():
    if 'user_id' not in session:
        return redirect(url_for('login_page'))
    org = get_current_organization()
    admin = User.query.get(session['user_id'])
    if not org or admin.permissions != '*':
        return redirect(url_for('admin_panel', error='Только суперадмин может загрузить тестовые данные.'))
    if org.test_data_loaded:
        return redirect(url_for('admin_panel', error='Тестовые данные уже загружены.'))
    _seed_employees_and_data(org)
    org.test_data_loaded = True
    db.session.commit()
    return redirect(url_for('admin_panel'))

@app.route('/admin/delete_test_data', methods=['POST'])
def admin_delete_test_data():
    if 'user_id' not in session:
        return redirect(url_for('login_page'))
    org = get_current_organization()
    admin = User.query.get(session['user_id'])
    if not org or admin.permissions != '*':
        return redirect(url_for('admin_panel', error='Только суперадмин может удалить тестовые данные.'))
    users_to_delete = User.query.filter_by(organization_id=org.id, role='user').all()
    for u in users_to_delete:
        WorkSession.query.filter_by(user_id=u.id).delete()
        TaskCompletion.query.filter_by(user_id=u.id).delete()
        for t in list(u.tasks):
            t.users.remove(u)
        db.session.delete(u)
    Task.query.filter_by(organization_id=org.id).delete()
    org.test_data_loaded = False
    db.session.commit()
    return redirect(url_for('admin_panel'))

# ---------- Employee routes ----------
@app.route('/admin/add_user', methods=['POST'])
def add_user():
    org = get_current_organization()
    admin = User.query.get(session['user_id'])
    if not org or not has_perm(admin, 'profiles'):
        return redirect(url_for('admin_panel', error='Нет прав для управления сотрудниками.'))
    login = request.form['login']
    if User.query.filter_by(login=login, organization_id=org.id).first():
        return redirect(url_for('admin_panel', error=f"Логин '{login}' уже занят."))
    u = User(
        first_name=request.form['f_name'],
        last_name=request.form['l_name'],
        login=login,
        password=request.form['password'],
        hourly_rate=float(request.form['rate']),
        role='user',
        organization_id=org.id
    )
    db.session.add(u)
    db.session.commit()
    return redirect(url_for('admin_panel'))

@app.route('/admin/edit_user/<int:uid>', methods=['POST'])
def edit_user(uid):
    org = get_current_organization()
    admin = User.query.get(session['user_id'])
    if not org or not has_perm(admin, 'profiles'):
        return redirect(url_for('admin_panel', error='Нет прав для управления сотрудниками.'))
    u = User.query.get_or_404(uid)
    if u.organization_id != org.id:
        return "Forbidden", 403
    new_login = request.form['login']
    if new_login != u.login and User.query.filter_by(login=new_login, organization_id=org.id).first():
        return redirect(url_for('admin_panel', error=f"Логин '{new_login}' уже занят."))
    u.first_name = request.form['f_name']
    u.last_name = request.form['l_name']
    u.login = new_login
    if request.form.get('password'):
        u.password = request.form['password']
    u.hourly_rate = float(request.form['rate'])
    db.session.commit()
    return redirect(url_for('admin_panel'))

@app.route('/admin/delete_user/<int:uid>', methods=['POST'])
def delete_user(uid):
    org = get_current_organization()
    admin = User.query.get(session['user_id'])
    if not org or not has_perm(admin, 'profiles'):
        return redirect(url_for('admin_panel', error='Нет прав для управления сотрудниками.'))
    u = User.query.get_or_404(uid)
    if u.organization_id != org.id:
        return "Forbidden", 403
    WorkSession.query.filter_by(user_id=uid).delete()
    TaskCompletion.query.filter_by(user_id=uid).delete()
    for t in list(u.tasks):
        t.users.remove(u)
    db.session.delete(u)
    db.session.commit()
    return redirect(url_for('admin_panel'))

# ---------- Admin management routes ----------
@app.route('/admin/add_admin', methods=['POST'])
def add_admin():
    org = get_current_organization()
    current = User.query.get(session['user_id'])
    if not org or not is_superadmin(current):
        return redirect(url_for('admin_panel', error='Только суперадмин может создавать администраторов.'))
    login = request.form['login']
    if User.query.filter_by(login=login, organization_id=org.id).first():
        return redirect(url_for('admin_panel', error=f"Логин '{login}' уже занят."))
    perms_selected = request.form.getlist('perms')
    permissions = ','.join(perms_selected) if perms_selected else ''
    u = User(
        first_name=request.form['f_name'],
        last_name=request.form['l_name'],
        login=login,
        password=request.form['password'],
        hourly_rate=0,
        role='admin',
        permissions=permissions,
        organization_id=org.id
    )
    db.session.add(u)
    db.session.commit()
    return redirect(url_for('admin_panel'))

@app.route('/admin/edit_admin/<int:uid>', methods=['POST'])
def edit_admin(uid):
    org = get_current_organization()
    current = User.query.get(session['user_id'])
    if not org or not is_superadmin(current):
        return redirect(url_for('admin_panel', error='Только суперадмин может редактировать администраторов.'))
    u = User.query.get_or_404(uid)
    if u.organization_id != org.id or u.role != 'admin':
        return "Forbidden", 403
    if u.permissions == '*':
        return redirect(url_for('admin_panel', error='Нельзя изменить суперадмина.'))
    new_login = request.form['login']
    if new_login != u.login and User.query.filter_by(login=new_login, organization_id=org.id).first():
        return redirect(url_for('admin_panel', error=f"Логин '{new_login}' уже занят."))
    u.first_name = request.form['f_name']
    u.last_name = request.form['l_name']
    u.login = new_login
    if request.form.get('password'):
        u.password = request.form['password']
    perms_selected = request.form.getlist('perms')
    u.permissions = ','.join(perms_selected) if perms_selected else ''
    db.session.commit()
    return redirect(url_for('admin_panel'))

@app.route('/admin/delete_admin/<int:uid>', methods=['POST'])
def delete_admin(uid):
    org = get_current_organization()
    current = User.query.get(session['user_id'])
    if not org or not is_superadmin(current):
        return redirect(url_for('admin_panel', error='Только суперадмин может удалять администраторов.'))
    u = User.query.get_or_404(uid)
    if u.organization_id != org.id or u.role != 'admin':
        return "Forbidden", 403
    if u.permissions == '*':
        return redirect(url_for('admin_panel', error='Нельзя удалить суперадмина.'))
    db.session.delete(u)
    db.session.commit()
    return redirect(url_for('admin_panel'))

# ---------- Task routes ----------
@app.route('/admin/add_task', methods=['POST'])
def add_task():
    org = get_current_organization()
    admin = User.query.get(session['user_id'])
    if not org or not has_perm(admin, 'tasks'):
        return redirect(url_for('admin_panel', error='Нет прав для управления задачами.'))
    dl = datetime.strptime(request.form['deadline'], '%Y-%m-%d').date()
    t = Task(title=request.form['title'], deadline=dl, organization_id=org.id)
    db.session.add(t)
    db.session.flush()
    for uid in request.form.getlist('user_ids'):
        u = User.query.get(int(uid))
        if u and u.organization_id == org.id:
            t.users.append(u)
            db.session.add(TaskCompletion(task_id=t.id, user_id=u.id, completed=False))
    db.session.commit()
    return redirect(url_for('admin_panel'))

@app.route('/admin/edit_task/<int:tid>', methods=['POST'])
def edit_task(tid):
    org = get_current_organization()
    admin = User.query.get(session['user_id'])
    if not org or not has_perm(admin, 'tasks'):
        return redirect(url_for('admin_panel', error='Нет прав для управления задачами.'))
    t = Task.query.get_or_404(tid)
    if t.organization_id != org.id:
        return "Forbidden", 403
    t.title = request.form['title']
    t.deadline = datetime.strptime(request.form['deadline'], '%Y-%m-%d').date()
    new_user_ids = set(int(uid) for uid in request.form.getlist('user_ids'))
    current_user_ids = {u.id for u in t.users}
    for uid in current_user_ids - new_user_ids:
        u = User.query.get(uid)
        t.users.remove(u)
        TaskCompletion.query.filter_by(task_id=t.id, user_id=uid).delete()
    for uid in new_user_ids - current_user_ids:
        u = User.query.get(uid)
        if u and u.organization_id == org.id:
            t.users.append(u)
            db.session.add(TaskCompletion(task_id=t.id, user_id=uid, completed=False))
    db.session.commit()
    return redirect(url_for('admin_panel'))

@app.route('/admin/delete_task/<int:tid>', methods=['POST'])
def delete_task(tid):
    org = get_current_organization()
    admin = User.query.get(session['user_id'])
    if not org or not has_perm(admin, 'tasks'):
        return redirect(url_for('admin_panel', error='Нет прав для управления задачами.'))
    t = Task.query.get_or_404(tid)
    if t.organization_id != org.id:
        return "Forbidden", 403
    TaskCompletion.query.filter_by(task_id=tid).delete()
    db.session.delete(t)
    db.session.commit()
    return redirect(url_for('admin_panel'))

@app.route('/task/done/<int:tid>')
def mark_done_web(tid):
    if 'user_id' not in session:
        return redirect(url_for('login_page'))
    user = User.query.get(session['user_id'])
    task = Task.query.get_or_404(tid)
    if user not in task.users:
        return "Forbidden", 403
    tc = TaskCompletion.query.filter_by(task_id=tid, user_id=user.id).first()
    if tc:
        tc.completed = True
        db.session.commit()
        total = len(task.users)
        completed_count = TaskCompletion.query.filter_by(task_id=tid, completed=True).count()
        if completed_count == total:
            task.is_done = True
            db.session.commit()
    return redirect(url_for('dashboard'))

@app.route('/admin/generate_invite', methods=['POST'])
def generate_invite():
    org = get_current_organization()
    admin = User.query.get(session['user_id'])
    if not org or admin.role != 'admin':
        return jsonify({"error": "forbidden"}), 403
    org.invite_code = generate_invite_code()
    org.invite_code_expires = datetime.utcnow() + timedelta(minutes=15)
    db.session.commit()
    expires_ts = int(org.invite_code_expires.timestamp() * 1000)
    return jsonify({"invite_code": org.invite_code, "expires": org.invite_code_expires.strftime('%H:%M'), "expires_ts": expires_ts})

# ---------- API: reveal password ----------
@app.route('/api/admin/reveal_password', methods=['POST'])
def reveal_password():
    if 'user_id' not in session:
        return jsonify({"error": "unauthorized"}), 401
    admin = User.query.get(session['user_id'])
    if admin.role != 'admin':
        return jsonify({"error": "forbidden"}), 403
    data = request.json
    user_id = data.get('user_id')
    admin_password = data.get('admin_password')
    if not admin_password or admin.password != admin_password:
        return jsonify({"error": "Неверный пароль администратора"}), 403
    org = get_current_organization()
    u = User.query.get(user_id)
    if not u or u.organization_id != org.id:
        return jsonify({"error": "not found"}), 404
    return jsonify({"password": u.password})

# ---------- API: Stats ----------
@app.route('/api/stats')
def get_stats():
    if 'user_id' not in session:
        return jsonify({"error": "unauthorized"}), 401
    requesting_user = User.query.get(session['user_id'])
    target_uid = request.args.get('target_uid')
    if target_uid and requesting_user.role == 'admin':
        org = get_current_organization()
        user = User.query.get(int(target_uid))
        if not user or user.organization_id != org.id:
            return jsonify({"error": "not found"}), 404
    else:
        user = requesting_user

    period = request.args.get('period', 'week')
    date_str = request.args.get('date')
    custom_start = request.args.get('start')
    custom_end = request.args.get('end')
    today = date.today()

    if date_str:
        try:
            day = datetime.strptime(date_str, '%Y-%m-%d').date()
            sessions = WorkSession.query.filter(WorkSession.user_id == user.id, WorkSession.date == day).all()
            total_min = sum(s.duration_minutes for s in sessions)
            return jsonify({"date": day.strftime('%d.%m.%Y'), "hours": round(total_min / 60, 2),
                            "earnings": round((total_min / 60) * user.hourly_rate, 2)})
        except:
            return jsonify({"error": "invalid date"}), 400

    start, end = get_period_bounds(period, custom_start, custom_end)
    sessions = WorkSession.query.filter(WorkSession.user_id == user.id, WorkSession.date >= start, WorkSession.date <= end).all()
    total_min = sum(s.duration_minutes for s in sessions)
    total_hours = total_min / 60
    earnings = total_hours * user.hourly_rate

    if (end - start).days > 90:
        labels, values = [], []
        current = start.replace(day=1)
        while current <= end:
            month_min = sum(s.duration_minutes for s in sessions if s.date.year == current.year and s.date.month == current.month)
            labels.append(current.strftime('%b %Y'))
            values.append(round(month_min / 60, 2))
            current = current.replace(month=current.month % 12 + 1, year=current.year + (1 if current.month == 12 else 0))
    else:
        labels, values = [], []
        current = start
        while current <= end:
            day_min = sum(s.duration_minutes for s in sessions if s.date == current)
            labels.append(current.strftime('%d.%m'))
            values.append(round(day_min / 60, 2))
            current += timedelta(days=1)

    return jsonify({"labels": labels, "values": values, "total_hours": round(total_hours, 2),
                    "earnings": round(earnings, 2), "rate": user.hourly_rate})

@app.route('/api/user/full_stats')
def user_full_stats():
    user_id = request.args.get('user_id')
    period = request.args.get('period', 'week')
    if not user_id:
        return jsonify({"error": "user_id required"}), 400
    u = User.query.get(int(user_id))
    if not u:
        return jsonify({"error": "not found"}), 404
    start, end = get_period_bounds(period)
    sessions = WorkSession.query.filter(WorkSession.user_id == u.id, WorkSession.date >= start, WorkSession.date <= end).all()
    total_min = sum(s.duration_minutes for s in sessions)
    total_hours = total_min / 60
    earnings = total_hours * u.hourly_rate

    if (end - start).days > 90:
        labels, values = [], []
        current = start.replace(day=1)
        while current <= end:
            month_min = sum(s.duration_minutes for s in sessions if s.date.year == current.year and s.date.month == current.month)
            labels.append(current.strftime('%b %Y'))
            values.append(round(month_min / 60, 2))
            current = current.replace(month=current.month % 12 + 1, year=current.year + (1 if current.month == 12 else 0))
    else:
        labels, values = [], []
        current = start
        while current <= end:
            day_min = sum(s.duration_minutes for s in sessions if s.date == current)
            labels.append(current.strftime('%d.%m'))
            values.append(round(day_min / 60, 2))
            current += timedelta(days=1)

    return jsonify({"labels": labels, "values": values, "total_hours": round(total_hours, 2),
                    "earnings": round(earnings, 2), "rate": u.hourly_rate})

@app.route('/api/admin/all_stats')
def admin_all_stats():
    if 'user_id' not in session:
        return jsonify({"error": "unauthorized"}), 401
    admin = User.query.get(session['user_id'])
    if admin.role != 'admin':
        return jsonify({"error": "forbidden"}), 403
    org = get_current_organization()
    users = User.query.filter_by(organization_id=org.id, role='user').all()
    period = request.args.get('period', 'week')
    custom_start = request.args.get('start')
    custom_end = request.args.get('end')
    today = date.today()
    start, end = get_period_bounds(period, custom_start, custom_end)

    result = []
    for u in users:
        sessions = WorkSession.query.filter(WorkSession.user_id == u.id, WorkSession.date >= start, WorkSession.date <= end).all()
        total_min = sum(s.duration_minutes for s in sessions)
        total_hours = total_min / 60
        earnings = total_hours * u.hourly_rate

        if (end - start).days > 90:
            labels, values = [], []
            current = start.replace(day=1)
            while current <= end:
                month_min = sum(s.duration_minutes for s in sessions if s.date.year == current.year and s.date.month == current.month)
                labels.append(current.strftime('%b %Y'))
                values.append(round(month_min / 60, 2))
                current = current.replace(month=current.month % 12 + 1, year=current.year + (1 if current.month == 12 else 0))
        else:
            labels, values = [], []
            current = start
            while current <= end:
                day_min = sum(s.duration_minutes for s in sessions if s.date == current)
                labels.append(current.strftime('%d.%m'))
                values.append(round(day_min / 60, 2))
                current += timedelta(days=1)

        result.append({"id": u.id, "name": f"{u.first_name} {u.last_name}", "labels": labels,
                       "values": values, "total_hours": round(total_hours, 2), "earnings": round(earnings, 2)})
    return jsonify(result)

# ---------- Client API ----------
@app.route('/api/check_invite', methods=['POST'])
def check_invite():
    data = request.json
    code = data.get('invite_code')
    org = Organization.query.filter_by(invite_code=code).first()
    if org and org.invite_code_expires and org.invite_code_expires > datetime.utcnow():
        return jsonify({"valid": True, "org_id": org.id, "org_name": org.name})
    return jsonify({"valid": False}), 400

@app.route('/api/admin/verify', methods=['POST'])
def api_admin_verify():
    d = request.json or {}
    org_id = d.get('org_id')
    login = d.get('login')
    password = d.get('password')
    if not org_id or not login or not password:
        return jsonify({"error": "missing data"}), 400
    org = Organization.query.get(org_id)
    if not org:
        return jsonify({"error": "Invalid organization"}), 401
    u = User.query.filter_by(login=login, password=password, organization_id=org.id, role='admin').first()
    if u:
        return jsonify({"id": u.id, "name": f"{u.first_name} {u.last_name}",
                        "role": "admin", "org_id": org.id})
    return jsonify({"error": "Неверные данные или не администратор"}), 403

@app.route('/api/login', methods=['POST'])
def api_login():
    d = request.json
    org_id = d.get('org_id')
    login = d.get('login')
    password = d.get('password')
    if not org_id or not login or not password:
        return jsonify({"error": "missing data"}), 400
    org = Organization.query.get(org_id)
    if not org:
        return jsonify({"error": "Invalid organization"}), 401
    u = User.query.filter_by(login=login, password=password, organization_id=org.id).first()
    if u:
        if u.role == 'admin':
            return jsonify({"error": "Администраторы не могут входить через клиент"}), 403
        tasks = []
        for t in u.tasks:
            total = len(t.users)
            completed = TaskCompletion.query.filter_by(task_id=t.id, completed=True).count()
            tasks.append({"id": t.id, "title": t.title, "deadline": str(t.deadline),
                          "urgent": (t.deadline - date.today()).days <= 2 and not t.is_done,
                          "assignees": [f"{user.first_name} {user.last_name}" for user in t.users],
                          "completed_by_me": TaskCompletion.query.filter_by(task_id=t.id, user_id=u.id, completed=True).first() is not None,
                          "progress": f"{completed}/{total}", "is_done": t.is_done})
        hist = WorkSession.query.filter_by(user_id=u.id).order_by(WorkSession.date.desc()).limit(5).all()
        today_sessions = WorkSession.query.filter_by(user_id=u.id, date=date.today()).all()
        today_min = sum(s.duration_minutes for s in today_sessions)
        return jsonify({"id": u.id, "name": u.first_name, "rate": u.hourly_rate, "role": u.role,
                        "today_min": today_min,
                        "tasks": tasks,
                        "chart": {"labels": [s.date.strftime('%d.%m') for s in reversed(hist)],
                                  "values": [round(s.duration_minutes / 60, 2) for s in reversed(hist)]}})
    return jsonify({"error": "Invalid credentials"}), 401

@app.route('/api/sync', methods=['POST'])
def sync_time():
    d = request.json
    ws = WorkSession.query.filter_by(user_id=d['user_id'], date=date.today()).first()
    if not ws:
        ws = WorkSession(user_id=d['user_id'])
        db.session.add(ws)
    ws.duration_minutes += 1
    db.session.commit()
    return jsonify({"status": "ok"})

@app.route('/api/tasks', methods=['GET'])
def get_tasks_api():
    user_id = request.args.get('user_id')
    if not user_id:
        return jsonify({"error": "user_id required"}), 400
    u = User.query.get(int(user_id))
    if not u:
        return jsonify({"error": "user not found"}), 404
    tasks = []
    for t in u.tasks:
        total = len(t.users)
        completed = TaskCompletion.query.filter_by(task_id=t.id, completed=True).count()
        tasks.append({"id": t.id, "title": t.title, "deadline": str(t.deadline), "is_done": t.is_done,
                      "urgent": (t.deadline - date.today()).days <= 2 and not t.is_done,
                      "assignees": [f"{user.first_name} {user.last_name}" for user in t.users],
                      "completed_by_me": TaskCompletion.query.filter_by(task_id=t.id, user_id=u.id, completed=True).first() is not None,
                      "progress": f"{completed}/{total}"})
    return jsonify(tasks)

@app.route('/api/task/<int:task_id>/done', methods=['POST'])
def mark_task_done_api(task_id):
    d = request.json or {}
    user_id = d.get('user_id')
    if not user_id:
        return jsonify({"error": "user_id required"}), 400
    u = User.query.get(int(user_id))
    task = Task.query.get_or_404(task_id)
    if u not in task.users:
        return jsonify({"error": "forbidden"}), 403
    tc = TaskCompletion.query.filter_by(task_id=task_id, user_id=u.id).first()
    if tc:
        tc.completed = True
        db.session.commit()
        total = len(task.users)
        if TaskCompletion.query.filter_by(task_id=task_id, completed=True).count() == total:
            task.is_done = True
            db.session.commit()
        return jsonify({"status": "ok"})
    return jsonify({"error": "completion record not found"}), 404

@app.route('/api/admin/adjust_time_web', methods=['POST'])
def admin_adjust_time_web():
    if 'user_id' not in session:
        return jsonify({"error": "unauthorized"}), 401
    admin = User.query.get(session['user_id'])
    if not admin or admin.role != 'admin':
        return jsonify({"error": "forbidden"}), 403
    data = request.json or {}
    user_id = data.get('user_id')
    minutes = data.get('minutes')
    adj_date_str = data.get('date')
    if not user_id or minutes is None:
        return jsonify({"error": "missing data"}), 400
    u = User.query.get(user_id)
    if not u or u.organization_id != admin.organization_id:
        return jsonify({"error": "user not found"}), 404
    try:
        adj_date = date.fromisoformat(adj_date_str) if adj_date_str else date.today()
    except Exception:
        adj_date = date.today()
    ws = WorkSession.query.filter_by(user_id=user_id, date=adj_date).first()
    if not ws:
        ws = WorkSession(user_id=user_id, date=adj_date, duration_minutes=0)
        db.session.add(ws)
    new_duration = max(0, ws.duration_minutes + minutes)
    ws.duration_minutes = new_duration
    ws.manual_adjustment = True
    ws.adjusted_by = admin.id
    db.session.commit()
    return jsonify({"status": "ok", "new_total": ws.duration_minutes})

@app.route('/api/admin/delete_adjustment/<int:sid>', methods=['POST', 'DELETE'])
def admin_delete_adjustment(sid):
    if 'user_id' in session:
        admin = User.query.get(session['user_id'])
        if admin and admin.role == 'admin':
            ws = WorkSession.query.get(sid)
            if ws and ws.manual_adjustment:
                db.session.delete(ws)
                db.session.commit()
            return jsonify({"status": "ok"})
    data = request.json or {}
    admin_login = data.get('admin_login')
    admin_password = data.get('admin_password')
    if not admin_login or not admin_password:
        return jsonify({"error": "missing data"}), 400
    admin = User.query.filter_by(login=admin_login, password=admin_password).first()
    if not admin or admin.role != 'admin':
        return jsonify({"error": "forbidden"}), 403
    ws = WorkSession.query.get(sid)
    if not ws or ws.adjusted_by != admin.id:
        return jsonify({"error": "not found"}), 404
    if ws.manual_adjustment:
        db.session.delete(ws)
        db.session.commit()
    return jsonify({"status": "ok"})

@app.route('/api/admin/user_sessions/<int:uid>', methods=['GET', 'POST'])
def admin_user_sessions(uid):
    if request.method == 'GET':
        if 'user_id' not in session:
            return jsonify({"error": "unauthorized"}), 401
        admin = User.query.get(session['user_id'])
        if not admin or admin.role != 'admin':
            return jsonify({"error": "forbidden"}), 403
    else:
        data = request.json
        admin_login = data.get('admin_login')
        admin_password = data.get('admin_password')
        if not admin_login or not admin_password:
            return jsonify({"error": "missing data"}), 400
        admin = User.query.filter_by(login=admin_login, password=admin_password).first()
        if not admin or admin.role != 'admin':
            return jsonify({"error": "forbidden"}), 403

    u = User.query.get(uid)
    if not u or u.organization_id != admin.organization_id:
        return jsonify({"error": "user not found"}), 404
    sessions = WorkSession.query.filter_by(user_id=uid).order_by(WorkSession.date.desc()).limit(30).all()
    return jsonify([{"id": s.id, "date": s.date.strftime('%d.%m.%Y'), "minutes": s.duration_minutes,
                     "manual": s.manual_adjustment, "reason": s.adjust_reason} for s in sessions])

@app.route('/api/admin/users', methods=['POST'])
def admin_users():
    data = request.json
    org_id = data.get('org_id')
    login = data.get('login')
    password = data.get('password')
    if not org_id or not login or not password:
        return jsonify({"error": "missing data"}), 400
    org = Organization.query.get(org_id)
    if not org:
        return jsonify({"error": "Invalid organization"}), 401
    admin = User.query.filter_by(login=login, password=password, organization_id=org.id, role='admin').first()
    if not admin:
        return jsonify({"error": "Forbidden"}), 403
    users = User.query.filter_by(organization_id=org.id, role='user').all()
    return jsonify([{"id": u.id, "name": f"{u.first_name} {u.last_name}", "rate": u.hourly_rate} for u in users])

@app.route('/api/total_minutes')
def total_minutes():
    user_id = request.args.get('user_id')
    if not user_id:
        return jsonify({"error": "user_id required"}), 400
    sessions = WorkSession.query.filter_by(user_id=user_id, date=date.today()).all()
    return jsonify({"total_min": sum(s.duration_minutes for s in sessions)})

@app.route('/api/admin/report')
def download_report():
    if 'user_id' not in session:
        return jsonify({"error": "unauthorized"}), 401
    admin = User.query.get(session['user_id'])
    if admin.role != 'admin':
        return jsonify({"error": "forbidden"}), 403
    org = get_current_organization()
    period = request.args.get('period', 'month')
    month_str = request.args.get('month')
    custom_start = request.args.get('start')
    custom_end = request.args.get('end')

    if period == 'month' and month_str:
        try:
            year, month = map(int, month_str.split('-'))
            start_date = date(year, month, 1)
            end_date = date(year, month, monthrange(year, month)[1])
        except:
            start_date, end_date = get_period_bounds('month')
    else:
        start_date, end_date = get_period_bounds(period, custom_start, custom_end)

    users = User.query.filter_by(organization_id=org.id, role='user').all()
    wb = Workbook()
    ws = wb.active
    ws.title = f"Сводка {start_date.strftime('%d.%m.%Y')}–{end_date.strftime('%d.%m.%Y')}"
    headers = ['ID', 'Сотрудник', 'Ставка (₽/ч)', 'Отработано часов', 'Заработано (₽)']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    total_hours_sum = 0
    total_earnings_sum = 0
    for u in users:
        sessions = WorkSession.query.filter(WorkSession.user_id == u.id,
                                            WorkSession.date >= start_date, WorkSession.date <= end_date).all()
        total_min = sum(s.duration_minutes for s in sessions)
        hours = total_min / 60
        earnings = hours * u.hourly_rate
        total_hours_sum += hours
        total_earnings_sum += earnings
        ws.append([u.id, f"{u.first_name} {u.last_name}", u.hourly_rate, round(hours, 2), round(earnings, 2)])

    ws.append(['', 'ИТОГО', '', round(total_hours_sum, 2), round(total_earnings_sum, 2)])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for column in ws.columns:
        max_length = max((len(str(cell.value or '')) for cell in column), default=0)
        ws.column_dimensions[column[0].column_letter].width = max_length + 4

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"report_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        run_migrations()
    app.run(host='0.0.0.0', port=5000, debug=True)
